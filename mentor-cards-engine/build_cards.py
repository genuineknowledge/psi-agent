#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成 Mentor 周期报表卡 + Boss 全公司总览卡（飞书 Legacy 卡片 JSON）。

设计依据
--------
- 契约文档：``mentor-report-card-structure.md``（7 栏目规格 + 口径）
- 数据源：
  1. 《TODO LIST》飞书电子表格（wiki 挂载，地址见 data/sources.json，海豚解析写入）—— 填报/目标/逾期/趋势
  2. 飞书审批（请假）``data/leave.json`` —— 由 ``fetch_leave_attendance.py`` 拉取，⑤ 请假标注 + 豁免
  3. 飞书考勤（打卡）``data/attendance.json`` —— 由 ``fetch_leave_attendance.py`` 拉取，① 考勤异常
  4. 飞书通讯录（入职时间）``data/join.json`` —— 由 ``fetch_leave_attendance.py`` 拉取，
     **「未按时」只标入职后、非请假期间没写的人**：未入职 📅 与请假 🏖 都不计未按时
- 每张卡 = 7 栏目：① 人员概况 ② 目标数量 ③ 完成情况 ④ 逾期明细
  ⑤ 请假标注 ⑥ 评价概况 ⑦ 趋势；底部带数据源可点链接。
- Boss 卡 = 全局指标 + 团队维度 + **台账总览（各组 ②③⑥，尽可能多给出台账）** + 全公司逾期明细 + 请假 + 趋势 + 链接。

用法（海豚调用路径）
--------------------
1. 拉取真实请假/考勤（每次发卡前）：
   ``python3 mentor_cards/fetch_leave_attendance.py``   # 依赖 PSI_FEISHU_APP_ID/SECRET
2. 刷新 TODO LIST（数据源更新后）：
   ``feishu_doc_export`` 导出《TODO LIST》为 xlsx → ``python3 mentor_cards/build_cards.py --xlsx todo_list_source.xlsx``
3. 仅重新生成卡片（数据未变）：
   ``python3 mentor_cards/build_cards.py``
4. 读取产物发卡：``mentor_cards/mentor_cards.json``
   ``{mentor名: {"oid": open_id, "card": 卡片JSON}, ..., "__boss__": {"card": 卡片JSON}}``

口径说明（诚实标注）
--------------------
- 填报率 / 未按时 / 趋势：结构性事实，由 TODO LIST 自动计算。
- **未按时（⚠️）= 入职后 + 非请假 + 本周期没写**：
  ① 已批准请假（近周期窗口重叠）→ 请假豁免 🏖，不计未按时；
  ② 入职日期晚于本周期日（还没入职，如预填 08-31/09-02 的待入职）→ 📅 不计未按时；
  ③ 其余本周期没写 → ⚠️ 未按时。入职时间来自飞书通讯录 ``join_time``
  （``data/join.json``），后台未填过入职日期的默认 2026-03-01 不影响判定。
- 趋势填报率分母只计「该周期日之前已入职」的成员：中途入职的人（如 8.20 入职）
  在入职前的周期里没写是正常的，不进分母（如 7.24 期分母 31→25）。
- 请假：**飞书审批真实数据**（近 90 天，APPROVED 且与 8.21–8.28 窗口重叠）；
  未填报但处于已批准请假 → 记为「请假」豁免，不计未按时（如董修奇 8.03–9.01 暑假）。
- 考勤：**飞书打卡真实数据**（8.22–8.28），异常 = 迟到/早退/缺卡（NoNeedCheck/请假日不计）。
- 目标数量 / 完成情况：台账 ledger（层级/状态字段）在 **mentor 已使用台账**
  （填写评语/打分/状态）后接管；否则用**人工核对档案**
  （``data/manual_calibration.json``，卡面标注核对时间），不再有代码内数字。
- 逾期明细：台账 ledger（状态=逾期，mentor 已使用后）→ 关键词自动扫描补充
  （标注「自动标记」）→ 人工核对档案（``data/manual_calibration.json``）；
  请假豁免人员的未填报条目**不展示**（请假即豁免，见 ``_visible_overdue``），但真实请假信息
  保留在 ⑤ 请假标注（不含审批链接，普通成员无权限查看审批实例）。
- 评价概况：TODO LIST 无评分列，需独立评价台账（尚未接入）→ 显示「暂无评价数据」。

输出：``mentor_cards/mentor_cards.json``
"""

from __future__ import annotations

import argparse
import datetime
import json
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

# --------------------------------------------------------------------------
# 数据源地址（代码内不硬编码任何链接/token；全部由海豚运行时解析写入配置）：
#   · TODO LIST 文档 → data/sources.json["todo_list"]（海豚发卡前用飞书工具解析写入）
#   · 各 mentor 台账多维表格 → data/ledger_sources.json（海豚解析台账链接后填入）
# 各 mentor 台账多维表格来源清单（ledger_sources.json）：每个 mentor 一个台账 base
# （如 TODO 台账-孙逊）。台账是 ②目标数量 / ③完成情况 / ④逾期明细 / ⑥评价概况 的
# 权威来源。ledger_sources.json 只是「从哪取数」的地址清单（配置，不是数据）；
# 每次构建卡片时 load_ledger() 现场从飞书现取最新记录（不读预存快照），并顺手
# 落盘 data/ledger_<mentor>.json 作审计副本 / 断网回退。未注册台账的组回落
# 人工核对档案（data/manual_calibration.json），卡面标注口径。
# 卡片底部「报表」行 = 该 mentor 自己的台账链接（ledger_sources.json 里的 url）。

# 趋势取最近 N 个周期列（规格，不是数据）
TREND_WINDOW = 8

# --------------------------------------------------------------------------
# 人工核对档案（数据来源之一，存在 data/manual_calibration.json 而非代码里：
# 由海豚按《真知团队信息档案》逐周期核对后更新该文件；台账由 mentor 开始填写
# （评语/打分/状态）后自动接管 ②③④，见 ledger_active 闸门）
# --------------------------------------------------------------------------
def _load_mentor_oids_fallback() -> dict:
    """mentor 收卡 open_id 兜底表（data/mentor_oids.json，海豚用通讯录/roster 解析后维护）。"""
    data = _load_data("mentor_oids.json")
    if not isinstance(data, dict):
        return {}
    return {k: v for k, v in data.items() if not str(k).startswith("_")}


def load_calibration() -> dict:
    """读人工核对档案 data/manual_calibration.json（缺失时返回空结构）。"""
    p = DATA_DIR / "manual_calibration.json"
    if not p.exists():
        return {"goals": {}, "closure": {}, "overdue": {}, "calibrated_at": ""}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except (ValueError, OSError):
        return {"goals": {}, "closure": {}, "overdue": {}, "calibrated_at": ""}

# --------------------------------------------------------------------------
# 真实数据通道（fetch_leave_attendance.py 拉取；缺文件时优雅降级到兜底口径）
# --------------------------------------------------------------------------
HERE = Path(__file__).resolve().parent
DATA_DIR = HERE / "data"


def _load_data(name: str) -> list | dict | None:
    p = DATA_DIR / name
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except (ValueError, OSError):
        return None


_TODO_SRC_CACHE: dict | None = None


def load_todo_source(force: bool = False) -> dict:
    """TODO LIST 文档地址（data/sources.json，海豚每次发卡前用飞书工具解析后写入）。

    代码内不硬编码文档链接/token；缺失字段返回空字符串，调用方降级
    （不产出坏链接）。force=True 强制重新读盘。
    """
    global _TODO_SRC_CACHE
    if force or _TODO_SRC_CACHE is None:
        data = _load_data("sources.json")
        tl = ((data or {}).get("todo_list") or {}) if isinstance(data, dict) else {}
        _TODO_SRC_CACHE = {
            "name": tl.get("name", "TODO LIST"),
            "url": tl.get("url", ""),
            "obj_token": tl.get("obj_token", ""),
            "sheet_id": tl.get("sheet_id", ""),
            "doc_base": data.get("doc_base", "") if isinstance(data, dict) else "",
        }
    return _TODO_SRC_CACHE


def _todo_source_line() -> str:
    """卡片「数据源」行：有链接给可点链接，缺失时降级为纯文本（不产坏链接）。"""
    src = load_todo_source()
    if src["url"]:
        return f"📎 数据源：[打开 {src['name']} 电子表格]({src['url']})"
    return f"📎 数据源：{src['name']}（链接待海豚解析）"


LEAVE_DATA = _load_data("leave.json")
ATT_DATA = _load_data("attendance.json")
JOIN_DATA = _load_data("join.json")

YEAR = datetime.date.today().year  # TODO LIST 周期列无年份，按运行年份解析


def load_roster() -> dict[str, str]:
    """通讯录（fetch_leave_attendance.py 从飞书拉取）→ name → open_id。"""
    p = HERE / "roster.json"
    if not p.exists():
        return {}
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
        return {m["name"]: (m.get("open_id") or "") for m in data.get("members", [])}
    except (ValueError, OSError):
        return {}


def mentor_oids(names: list[str] | None = None) -> dict[str, str]:
    """mentor 收卡 open_id：优先通讯录 roster 按名解析；缺失回落到兜底表。

    names 为 None 时按兜底表名单解析（向后兼容：老调用方不传参行为不变）；
    传入由 TODO LIST 数据派生的 mentor 名单时覆盖全部 mentor——新 mentor
    只要在 roster.json（fetch_leave_attendance.py 产物）里就能拿到 open_id，
    不再被兜底表写死；roster 与兜底表都没有的人才缺失（调用方跳过发卡并告警）。
    """
    roster = load_roster()
    fallback = _load_mentor_oids_fallback()
    want = names if names is not None else list(fallback)
    if not roster:
        print("[warn] 未找到 roster.json（fetch_leave_attendance.py 产物），"
              "mentor open_id 仅能回落到兜底表；新 mentor 将缺失 open_id",
              file=sys.stderr)
    out: dict[str, str] = {}
    missing: list[str] = []
    for m in want:
        oid = (roster.get(m) or "") if roster else ""
        if not oid:
            oid = fallback.get(m, "")
        if not oid:
            missing.append(m)
        out[m] = oid
    if missing:
        print(f"[warn] mentor 无 open_id（roster 与兜底表均无）：{missing}，将跳过发卡",
              file=sys.stderr)
    return out


def runtime_windows(latest: str) -> tuple[tuple[str, str], str]:
    """由最新周期列推导 (请假豁免/展示窗口, 考勤统计窗口文本)。

    - 请假窗口：周期日往前 7 天 ~ 周期日（如 8.28 → 2026-08-21 ~ 2026-08-28）
    - 考勤窗口：周期日往前 6 天 ~ 周期日（与 fetch_leave_attendance.py 的 --cycle 一致）
    """
    cd = _cycle_date(latest)
    if not cd:
        return (("", ""), "")
    d = datetime.date.fromisoformat(cd)
    w0 = (d - datetime.timedelta(days=7)).isoformat()
    att_from = d - datetime.timedelta(days=6)
    att_text = f"{att_from.month}.{att_from.day}–{d.month}.{d.day}"
    return ((w0, cd), att_text)


def data_as_of() -> str:
    """数据拉取时间：取 leave/attendance/join 三个数据文件里最新的 fetched_at。"""
    ts = ""
    for name in ("attendance.json", "leave.json", "join.json"):
        data = _load_data(name)
        if isinstance(data, dict):
            t = data.get("fetched_at") or ""
            if t > ts:
                ts = t
    if ts:
        try:
            return datetime.datetime.fromisoformat(ts).strftime("%m-%d %H:%M")
        except ValueError:
            return ts[5:16]
    return "（未拉取）"


def load_ledger_sources() -> dict:
    """读 ledger_sources.json → {mentor: {name,url,app_token,tables}}（去掉 _ 开头的键）。"""
    p = HERE / "ledger_sources.json"
    if not p.exists():
        return {}
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
        return {k: v for k, v in data.items()
                if k and not k.startswith("_") and isinstance(v, dict)}
    except (ValueError, OSError):
        return {}


# 台账实时拉取：每次构建进程首次访问某 mentor 时现场从飞书现取（不读预存快照）。
# 进程内缓存避免同一进程重复请求；每次发卡脚本都是新进程 → 天然每次发卡现取。
_ledger_cache: dict[str, dict | None] = {}
_ledger_token: str | None = None


def _ledger_token_or_none() -> str | None:
    """懒拿 tenant token；环境变量缺失/取失败返回 None（调用方回退快照，不退出）。"""
    global _ledger_token
    if _ledger_token is not None:
        return _ledger_token
    if not (os.environ.get("PSI_FEISHU_APP_ID") and os.environ.get("PSI_FEISHU_APP_SECRET")):
        print("[warn] 缺 PSI_FEISHU_APP_ID / PSI_FEISHU_APP_SECRET，台账无法现取，"
              "回退落盘快照", file=sys.stderr)
        return None
    try:
        import fetch_ledgers
        _ledger_token = fetch_ledgers.get_tenant_token()
    except (Exception, SystemExit) as e:
        print(f"[warn] 拿飞书 tenant token 失败（{e}），台账回退落盘快照", file=sys.stderr)
        _ledger_token = None
    return _ledger_token


def load_ledger(mentor: str) -> dict | None:
    """该 mentor 台账 = **现场从飞书现取的最新记录**（ledger_sources.json 只是地址清单）。

    顺序：① 已注册台账 → 实时拉取（fetch_ledgers.fetch_ledger_for，顺手落盘审计副本）；
    ② 现取失败/未注册 → 回退落盘快照 data/ledger_<mentor>.json（断网兜底，打 warn）；
    ③ 都没有 → None（卡片回落人工核对档案 data/manual_calibration.json）。

    结构：{fetched_at, source{name,url,app_token,tables}, latest_cycle, cycles, rows[]}
    rows 每行 = {cycle, owners, mentors, level, level_raw, parent, title, due,
                 status, five, score, agent_score, comment, external,
                 comparison, guid, record_id, table}
    """
    if mentor in _ledger_cache:
        return _ledger_cache[mentor]

    src = load_ledger_sources().get(mentor, {})
    if src.get("app_token") and src.get("tables"):
        token = _ledger_token_or_none()
        if token:
            try:
                import fetch_ledgers
                ledger = fetch_ledgers.fetch_ledger_for(mentor, src, token, save=True)
                _ledger_cache[mentor] = ledger
                return ledger
            except Exception as e:
                print(f"[warn] {mentor} 台账现取失败（{e.__class__.__name__}: {e}），"
                      f"回退落盘快照", file=sys.stderr)
        # 未注册或现取失败 → 快照兜底
        p = DATA_DIR / f"ledger_{mentor}.json"
        if p.exists():
            try:
                ledger = json.loads(p.read_text(encoding="utf-8"))
                _ledger_cache[mentor] = ledger
                return ledger
            except (ValueError, OSError):
                pass
    _ledger_cache[mentor] = None
    return None


def ledger_latest_rows(ledger: dict | None) -> list[dict]:
    """台账最新周期行（去重：同标题保留评审过的行——有打分/评语 或 状态=已交付/进行中）。

    台账「周期日期」字段 = 每条任务归属周期；最新周期 = 台账当前周期。
    """
    if not ledger:
        return []
    rows = ledger.get("rows") or []
    latest = ledger.get("latest_cycle") or ""
    rows = [r for r in rows if r.get("cycle") == latest] if latest else rows
    seen: dict[str, dict] = {}
    for r in rows:
        key = (r.get("title") or "").strip()
        if not key:
            continue
        if key not in seen:
            seen[key] = r
            continue
        cur, new = seen[key], r

        def review_rank(x: dict) -> tuple[int, int]:
            return ((1 if (x.get("score") is not None or x.get("comment")) else 0),
                    (1 if x.get("status") in ("已交付", "进行中") else 0))
        if review_rank(new) > review_rank(cur):
            seen[key] = new
    return list(seen.values())


def goal_counts_from_ledger(rows: list[dict]) -> tuple[int, int, int] | None:
    """台账派生 (大目标, 小目标, TODO)：按 层级大类 计数（level 已归一化）。"""
    counts = {"大目标": 0, "小目标": 0, "TODO": 0}
    for r in rows:
        level = (r.get("level") or "").strip()
        if level in counts:
            counts[level] += 1
    if sum(counts.values()) == 0:
        return None
    return (counts["大目标"], counts["小目标"], counts["TODO"])


def closure_from_ledger(rows: list[dict]) -> tuple[int, int, int, int, int] | None:
    """台账派生 (已闭环, 进行中, 待开始, 请假顺延, 逾期)。

    - 已闭环 = 状态「已交付」；进行中 = 「进行中」；待开始 = 「待开始」；顺延 = 「请假顺延」。
    - 逾期 = 状态「待开始」且 截止日期早于今天（超期未交付，台账推导）。
    """
    counts = {"已交付": 0, "进行中": 0, "待开始": 0, "请假顺延": 0}
    today = datetime.date.today().isoformat()
    overdue = 0
    for r in rows:
        st = (r.get("status") or "").strip()
        if st in counts:
            counts[st] += 1
        if st == "待开始" and r.get("due") and r["due"] < today:
            overdue += 1
    if sum(counts.values()) == 0 and overdue == 0:
        return None
    return (counts["已交付"], counts["进行中"], counts["待开始"],
            counts["请假顺延"], overdue)


def evaluation_from_ledger(rows: list[dict]) -> dict | None:
    """台账派生 ⑥ 评价概况：打分非空行 → {avg, dist, n, comments}。无打分返回 None。"""
    scores = [r["score"] for r in rows if r.get("score") is not None]
    if not scores:
        return None
    dist = {"5": 0, "4": 0, "3": 0, "le2": 0}
    for s in scores:
        if s >= 5:
            dist["5"] += 1
        elif s >= 4:
            dist["4"] += 1
        elif s >= 3:
            dist["3"] += 1
        else:
            dist["le2"] += 1
    comments = [r["comment"] for r in rows if (r.get("comment") or "").strip()]
    return {
        "avg": round(sum(scores) / len(scores), 1),
        "dist": dist,
        "n": len(scores),
        "comments": comments,
    }


def overdue_from_ledger(rows: list[dict]) -> list[tuple[str, str, str]]:
    """台账派生逾期明细：状态=待开始 且 截止日期早于今天 → (负责人, 标题, '台账·超期未交付')。"""
    today = datetime.date.today().isoformat()
    out = []
    for r in rows:
        if (r.get("status") or "").strip() != "待开始":
            continue
        if not r.get("due") or r["due"] >= today:
            continue
        owner = r.get("owners") or "（未指定）"
        title = (r.get("title") or "").strip()[:60]
        if not title:
            continue
        out.append((owner, title, f"台账·超期未交付（截止 {r['due'][5:]}）"))
    return out


def ledger_has_mentor_input(rows: list[dict]) -> bool:
    """台账是否已有 mentor 填写痕迹（打分/评语非空）——「台账被使用」的信号。"""
    return any(r.get("score") is not None or (r.get("comment") or "").strip()
               for r in rows)


def calibration_goals(cal: dict, mentor: str) -> tuple[int, int, int]:
    g = cal.get("goals", {}).get(mentor) or (0, 0, 0)
    return tuple(g)[:3]  # type: ignore[return-value]


def calibration_closure(cal: dict, mentor: str) -> tuple[int, int, int]:
    c = cal.get("closure", {}).get(mentor) or (0, 0, 0)
    return tuple(c)[:3]  # type: ignore[return-value]


def calibration_overdue(cal: dict, mentor: str) -> list[tuple[str, str, str]]:
    return [tuple(row) for row in cal.get("overdue", {}).get(mentor, [])]  # type: ignore[misc]


def join_dates() -> dict[str, str]:
    """name -> 'YYYY-MM-DD'（来自飞书通讯录 join_time；无数据返回空 dict）。"""
    if not JOIN_DATA:
        return {}
    return {n: (v.get("join_date") or "") for n, v in JOIN_DATA.get("people", {}).items()}


def _cycle_date(col: str) -> str:
    """周期列名（'7.24' / '8.10日'）→ ISO 日期 '2026-07-24'；解析不出返回 ''。"""
    m = re.match(r"(\d{1,2})\.(\d{1,2})", col)
    if not m:
        return ""
    return f"{YEAR}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"


def joined_by(join_map: dict[str, str], name: str, cycle_date: str) -> bool:
    """该人在 cycle_date 之前是否已入职（join_date <= cycle_date）。

    无入职数据 → 默认视为已入职（不误伤）；join_date 为空同理。
    入职日期为默认值 2026-03-01 的人（飞书后台默认）不影响判定。
    """
    jd = join_map.get(name, "")
    if not jd or not cycle_date:
        return True
    return jd <= cycle_date


def approved_leaves() -> list[dict]:
    """已批准的请假实例（含申请人姓名）。"""
    if not LEAVE_DATA:
        return []
    out = []
    for e in LEAVE_DATA.get("entries", []):
        if e.get("status") != "APPROVED":
            continue
        if not e.get("start") or not e.get("end"):
            continue
        out.append(e)
    return out


def _overlaps_window(start: str, end: str, w0: str, w1: str) -> bool:
    """请假区间 [start,end] 是否与窗口 [w0,w1] 有重叠（ISO 日期字符串比较）。"""
    s, e, a, b = start[:10], end[:10], w0, w1
    return s <= b and e >= a


def leave_exempt_names(window: tuple[str, str] | None = None) -> set[str]:
    """本周期窗口内处于已批准请假的人（未填报时豁免，不计未按时/逾期）。

    window 由调用方按最新周期推导（runtime_windows），None 时返回空集。
    """
    if not window:
        return set()
    return {e["name"] for e in approved_leaves()
            if _overlaps_window(e["start"], e["end"], window[0], window[1])}


def group_leaves(names: list[str], window: tuple[str, str] | None = None) -> list[dict]:
    """某组内、窗口内有已批准请假的人员条目（相邻请假合并展示）。

    合并时保留最早一条审批的 ``instance_code``，供 ⑤ 请假标注展示请假信息。
    window 由调用方按最新周期推导（runtime_windows），None 时返回空列表。
    """
    if not window:
        return []
    rows = [e for e in approved_leaves() if e["name"] in names
            and _overlaps_window(e["start"], e["end"], window[0], window[1])]
    # 同人相邻请假合并：start 取最早、end 取最晚、天数相加
    by_name: dict[str, dict] = {}
    for e in rows:
        acc = by_name.setdefault(e["name"], {
            "name": e["name"], "start": e["start"], "end": e["end"],
            "days": 0.0, "types": set(), "reasons": [], "codes": [],
        })
        acc["start"] = min(acc["start"], e["start"])
        acc["end"] = max(acc["end"], e["end"])
        try:
            acc["days"] += float(e.get("interval") or 0)
        except ValueError:
            pass
        acc["types"].add(e.get("leave_type") or "请假")
        if e.get("reason"):
            acc["reasons"].append(e["reason"])
        if e.get("instance_code"):
            acc["codes"].append(e["instance_code"])
    for acc in by_name.values():
        acc["types"] = "、".join(sorted(acc["types"]))
        acc["reason"] = (acc["reasons"][0] if acc["reasons"] else "")
        acc["instance_code"] = acc["codes"][0] if acc["codes"] else ""
    return list(by_name.values())


def _visible_overdue(rows: list[tuple[str, str, str]],
                     exempt: set[str]) -> list[tuple[str, str, str]]:
    """逾期明细过滤：请假豁免人员的「未填报」条目不展示（请假即豁免）。

    任务延期类条目（如高博的延期标记）不受影响；非豁免人员的未填报保留。
    """
    return [row for row in rows
            if not (row[0] in exempt and "未填报" in row[1])]


def attendance_anomalies(names: set[str]) -> dict[str, list[str]]:
    """8.22–8.28 打卡异常：{姓名: ["缺卡5", "迟到2"]}（NoNeedCheck/正常不计）。"""
    if not ATT_DATA:
        return {}
    label = {"Late": "迟到", "Early": "早退", "Lack": "缺卡"}
    counter: dict[str, dict[str, int]] = {}
    for r in ATT_DATA.get("results", []):
        name = r.get("name", "")
        if name not in names:
            continue
        for res in (r.get("check_in_result", ""), r.get("check_out_result", "")):
            if res in label:
                c = counter.setdefault(name, {})
                c[label[res]] = c.get(label[res], 0) + 1
    out: dict[str, list[str]] = {}
    for name, c in counter.items():
        items = sorted(c.items(), key=lambda kv: -kv[1])
        out[name] = [f"{k}{v}" for k, v in items]
    return out


# --------------------------------------------------------------------------
# 工具函数
# --------------------------------------------------------------------------
_DELAY_RE = re.compile(
    r"(delay\s*(自|到|至|ed| to)?|延期|延后|顺延|delayed|无限期|要 delay|看起来要 delay)",
    re.IGNORECASE,
)
# 正文里常见但**不是**延期标记的规划词
_NOT_DELAY = re.compile(r"年底|明年|12\s*月|年度")

# TODO LIST 表头里**不是**周期列的固定列名（其余列一律要求能解析出日期，
# 否则会被当成周期列 → 请假/考勤窗口全空 → 全员误标 ⚠️ 未按时）
_NON_CYCLE_COLS = {"任务负责人", "mentor", "test_read"}


def _is_cycle_col(name: str) -> bool:
    """周期列判定：非固定列名且列名能解析出日期（'7.24' / '8.10日'）。"""
    if not name or name in _NON_CYCLE_COLS:
        return False
    return bool(_cycle_date(name))


def _drop_unparsable_latest(cols: list[str]) -> list[str]:
    """最新周期列解析不出日期时告警并回退到上一个真实周期列（防呆）。

    正常情况下 load_people 已把 date_cols 过滤干净；此函数是第二道保险：
    解析不出日期的列被过滤后本不该出现，一旦出现（如旧版 parsed.json
    缓存混入坏列）就逐列回退，绝不默默把最新列当周期用导致全员红卡。
    """
    while cols and not _cycle_date(cols[-1]):
        print(f"[warn] 周期列 {cols[-1]!r} 解析不出日期，"
              "已回退到上一个真实周期列", file=sys.stderr)
        cols.pop()
    return cols


def load_people(xlsx: str | None) -> tuple[list[str], list[dict]]:
    """载入人员数据。

    --xlsx 给出时：用 openpyxl 重新解析飞书导出的表格 → 重写 todo_list_parsed.json。
    否则：直接读已解析的 todo_list_parsed.json。
    返回 (date_cols, people)。
    """
    if xlsx:
        from openpyxl import load_workbook

        wb = load_workbook(xlsx, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        col = {name: i for i, name in enumerate(header)}
        date_cols = _drop_unparsable_latest(
            [name for name in header if _is_cycle_col(name)])
        people = []
        for r in rows[1:]:
            name = ""
            if "任务负责人" in col and col["任务负责人"] < len(r) and r[col["任务负责人"]] is not None:
                name = str(r[col["任务负责人"]]).strip().replace("@", "")
            mentor = ""
            if "mentor" in col and col["mentor"] < len(r) and r[col["mentor"]] is not None:
                mentor = str(r[col["mentor"]]).strip().replace("@", "")
            if not name:
                continue
            cols = {}
            for c in date_cols:
                idx = col[c]
                cols[c] = str(r[idx]).strip() if idx < len(r) and r[idx] is not None else ""
            people.append({"name": name, "mentor": mentor, "cols": cols})
        out = {"date_cols": date_cols, "people": people}
        parsed_path = Path(__file__).resolve().parent.parent / "todo_list_parsed.json"
        parsed_path.write_text(
            json.dumps(out, ensure_ascii=False, indent=1), encoding="utf-8")
        print(f"[xlsx] 已重写 {parsed_path}（{len(people)} 人 / {len(date_cols)} 周期列）")
        return date_cols, people

    parsed = json.loads(
        (Path(__file__).resolve().parent.parent / "todo_list_parsed.json")
        .read_text(encoding="utf-8"))
    # 缓存文件同样过一遍防御：过滤非周期列 + 最新列解析不出时回退
    date_cols = _drop_unparsable_latest(
        [c for c in parsed["date_cols"] if _is_cycle_col(c)])
    return date_cols, parsed["people"]


def member_status(person: dict, latest: str, exempt: set[str],
                  join_map: dict[str, str] | None = None,
                  cycle_date: str | None = None) -> str:
    """该人本周期状态：'leave'（请假/放假）/ 'filled' / 'not_joined' / 'unfilled'。

    判定顺序（「入职后、非请假期间没写」才标未按时）：
      1. 正文含「放假」→ leave；
      2. 未填报 + 已批准请假（本周期窗口）→ leave（请假豁免）；
      3. 未填报 + 入职日期 > 本周期日（还没入职）→ not_joined（不应填，不计未按时）；
      4. 未填报 + 已入职 + 非请假 → unfilled（真·该写没写）；
      5. 有内容 → filled。
    """
    if join_map is None:
        join_map = {}
    if cycle_date is None:
        cycle_date = _cycle_date(latest)
    text = (person["cols"].get(latest) or "").strip()
    if "放假" in text:
        return "leave"
    if not text:
        if person["name"] in exempt:
            return "leave"
        if cycle_date and not joined_by(join_map, person["name"], cycle_date):
            return "not_joined"
        return "unfilled"
    return "filled"


_LEAVE_ONLY_RE = re.compile(
    r"^\s*(?:请假|休假|放假|年假|事假|病假|调休|婚假|产假|陪产假|丧假|休息|居家)"
    r"(?:\d+(?:\.\d+)?\s*(?:天|日|号|天假))?\s*[，。、,. ]*\s*$")


def _real_fill(text: str) -> bool:
    """趋势口径：格子里写了真实内容才算「已填报」。

    整格只写「请假 / 事假2天 / 放假」这类请假标记 ≠ 填报 TODO
    （如 2026-08-14 贺雅诗格内仅「请假」二字，旧逻辑误计为已填报，
    导致该期填报率虚高）。附请假/休假/调休等字数说明的整格同样不计。
    """
    s = text.strip()
    return bool(s) and not _LEAVE_ONLY_RE.match(s)


def _leave_count_on(cycle_date: str, names: set[str]) -> int:
    """周期日当天处于已批准请假的人数（只数给定成员集合；无日期/无数据返回 0）。

    判断口径：已批准请假区间 [start, end] 覆盖该周期日（含端点），即该日在假。
    与 group_leaves 的 LEAVE_WINDOW 不同——趋势按「每一期的当日」回看，不受
    展示窗口限制，保证历史各期的请假人数都是当时的真实在假人数。
    """
    if not cycle_date:
        return 0
    seen: set[str] = set()
    for e in approved_leaves():
        s, t = (e.get("start") or "")[:10], (e.get("end") or "")[:10]
        if e.get("name") in names and s and t and s <= cycle_date <= t:
            seen.add(e["name"])
    return len(seen)


def trend_series(members: list[dict], date_cols: list[str], n: int,
                 join_map: dict[str, str] | None = None) -> list[tuple[str, int, int, int]]:
    """近 n 个周期列填报率序列：[(列名, 已填人数, 已入职人数, 当日请假在假人数), ...]。

    分母只计「该周期日之前已入职」的成员：中途入职的人（如 8.20 入职）
    在入职前的周期里没写是正常的，不进分母，否则早期填报率被拉低。
    第 4 元素 = 该周期日在假人数（已批准请假区间覆盖当日），供趋势括号标注。
    """
    join_map = join_map or {}
    cols = date_cols[-n:]
    series = []
    for c in cols:
        cd = _cycle_date(c)
        eligible = [p for p in members if joined_by(join_map, p["name"], cd)]
        filled = sum(1 for p in eligible if _real_fill(p["cols"].get(c) or ""))
        leave_n = _leave_count_on(cd, {p["name"] for p in eligible})
        series.append((c, filled, len(eligible), leave_n))
    return series


def direction_arrow(series: list[tuple[str, int, int, int]]) -> str:
    """最新 vs 上一列：↑ 绿 / → 灰 / ↓ 红。"""
    if len(series) < 2:
        return ""
    a, b = series[-2][1], series[-1][1]
    if b > a:
        return "🟢↑"
    if b == a:
        return "⚪→"
    return "🔴↓"


def auto_overdue_extra(people_in_group: list[dict], latest: str,
                       known: list[tuple[str, str, str]],
                       exempt: set[str],
                       join_map: dict[str, str] | None = None) -> list[tuple[str, str, str]]:
    """关键词扫描补充逾期（校准表之外的发现，标注「自动标记」）。

    只扫本周期列正文；跳过含「年底/明年/12月」的规划行；校准表里已有
    该责任人的不再重复追加；请假豁免人员与未入职人员跳过。
    """
    join_map = join_map or {}
    cd = _cycle_date(latest)
    known_names = {row[0] for row in known}
    extra = []
    for p in people_in_group:
        if p["name"] in exempt or p["name"] in known_names:
            continue
        if cd and not joined_by(join_map, p["name"], cd):
            continue
        text = p["cols"].get(latest) or ""
        if not text:
            continue
        for line in text.splitlines():
            if _DELAY_RE.search(line) and not _NOT_DELAY.search(line):
                snippet = line.strip()[:60]
                if snippet:
                    extra.append((p["name"], snippet, "自动标记"))
                break  # 一人一条，避免刷屏
    return extra


def auto_todo_count(people_in_group: list[dict], latest: str,
                    join_map: dict[str, str] | None = None) -> int:
    """自动解析本周期 TODO 主条目数（仅用于与人工校准表交叉校验，不覆盖人工值）。

    对每人本周期列正文，取「TODO」段（排除出现在「TODO 按时按格式提交率」
    这类正文引用里的字样，要求整行是 TODO 标题），数行首数字主条目，
    排除 4.1 这类子项编号。口径（主条目）与人工「组内汇总」本就不同，
    例如郑淳人工 16（含子项任务行）vs 自动 4，属正常差异；但「人工 0 且
    自动 >0」一定是数据源漏数，必须在 main() 里告警。未入职人员跳过。
    """
    join_map = join_map or {}
    cd = _cycle_date(latest)
    total = 0
    for p in people_in_group:
        if cd and not joined_by(join_map, p["name"], cd):
            continue
        text = p["cols"].get(latest) or ""
        m = re.search(r"(?im)^\s*TODO[^\n]*\n", text)
        if not m:
            continue
        seg = text[m.end():]
        items = re.findall(r"(?m)^\s*(\d+)(?!\.\d)[、.]\s*\S", seg)
        total += len(items)
    return total


def evaluation_text(evaluation: dict | None, src: str) -> str:
    """⑥ 评价概况文本：台账有打分 → 平均分 + 分布 + 评语条数；否则「暂无评价数据」。"""
    if not evaluation:
        return "**⑥ 评价概况**\n暂无评价数据（台账未填写打分/评语）"
    d = evaluation["dist"]
    parts = [f"**⑥ 评价概况**（{src}）",
             f"平均 {evaluation['avg']} ★ · 已评 {evaluation['n']} 条"]
    bucket = []
    for label, n in (("5★", d["5"]), ("4★", d["4"]), ("3★", d["3"]), ("≤2★", d["le2"])):
        if n:
            bucket.append(f"{label}×{n}")
    if bucket:
        parts.append(" ｜ ".join(bucket))
    if evaluation.get("comments"):
        parts.append(f"评语 {len(evaluation['comments'])} 条")
    return "\n".join(parts)


def health_template(overdue_count: int, unfilled_count: int,
                    filled_ratio: float) -> str:
    """卡头配色：有逾期/未填 → red；全员填报且无逾期 → green；其余 blue。"""
    if overdue_count > 0 or unfilled_count > 0:
        return "red"
    if filled_ratio >= 1.0:
        return "green"
    return "blue"


def esc(text: str) -> str:
    """lark_md 转义：防正文中的 < > & 破坏 markdown。"""
    return (text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


def _md_div(content: str) -> dict:
    return {"tag": "div", "text": {"tag": "lark_md", "content": content}}


def _note(text: str) -> dict:
    return {"tag": "note", "elements": [{"tag": "plain_text", "content": text}]}


def todo_snapshot() -> str:
    """TODO LIST 解析时间：取 todo_list_parsed.json 的修改时间（导出/解析即当时）。"""
    p = Path(__file__).resolve().parent.parent / "todo_list_parsed.json"
    if p.exists():
        try:
            return datetime.datetime.fromtimestamp(p.stat().st_mtime).strftime("%m-%d %H:%M")
        except OSError:
            pass
    return ""


def source_note(as_of: str = "", att_window: str = "") -> str:
    src = load_todo_source()
    loc = f"{src['url']} · obj_token {src['obj_token']}" if src["url"] else "TODO 总表链接待海豚解析"
    return (
        f"数据源：{src['name']}（TODO 截至 {todo_snapshot() or '—'}）· "
        f"飞书审批（请假）· 飞书考勤（打卡 {att_window}，截至 {as_of}）\n"
        f"{loc}"
    )


# --------------------------------------------------------------------------
# 卡片构建
# --------------------------------------------------------------------------
def build_mentor_card(mentor: str, members: list[dict], latest: str,
                      date_cols: list[str], exempt: set[str],
                      join_map: dict[str, str]) -> dict:
    """一张 mentor 摘要卡（7 栏目 + 成员状态 + 考勤异常 + 数据源链接）。

    未按时（⚠️）只标「入职后、非请假期间没写 todo」的人：
    请假/放假 🏖 与未入职 📅 都不计未按时。
    """
    cd = _cycle_date(latest)
    leave_window, att_window = runtime_windows(latest)
    as_of = data_as_of()
    ledger = load_ledger(mentor)
    ledger_rows = ledger_latest_rows(ledger)
    ledger_src = load_ledger_sources().get(mentor, {})
    cal = load_calibration()
    total = len(members)
    filled = sum(1 for p in members if member_status(p, latest, exempt, join_map, cd) == "filled")
    leave_names = {e["name"] for e in group_leaves([p["name"] for p in members], leave_window)}
    leave = len(leave_names)
    unfilled = sum(1 for p in members if member_status(p, latest, exempt, join_map, cd) == "unfilled")
    unfilled_names = [p["name"] for p in members
                      if member_status(p, latest, exempt, join_map, cd) == "unfilled"]
    not_joined = sum(1 for p in members
                     if member_status(p, latest, exempt, join_map, cd) == "not_joined")

    # ② 目标数量 / ③ 完成情况 / ④ 逾期明细 / ⑥ 评价概况 —— 来源优先级：
    #   该 mentor 台账（fetch_ledgers.py 从 ledger_sources.json 现取，最新周期）→ 台账口径；
    #   否则 → 人工核对档案（data/manual_calibration.json，卡面标注核对时间）。
    if ledger_rows:
        goals = goal_counts_from_ledger(ledger_rows) or (0, 0, 0)
        closure = closure_from_ledger(ledger_rows) or (0, 0, 0, 0, 0)
        goals_src = f"台账·截至{ledger.get('latest_cycle', '')[-5:]}"
        known_overdue = overdue_from_ledger(ledger_rows)
    else:
        goals = calibration_goals(cal, mentor)
        c3 = calibration_closure(cal, mentor) or (0, 0, 0)
        closure = (c3[0], c3[1], 0, 0, c3[2])   # (已闭环, 进行中, 待开始0, 顺延0, 逾期)
        goals_src = f"人工核对 {cal.get('calibrated_at', '')[:10]}"
        known_overdue = calibration_overdue(cal, mentor)
    goals_b, goals_s, goals_t = goals
    closed, doing, waiting, deferred, overdue_cnt = closure
    evaluation = evaluation_from_ledger(ledger_rows) if ledger_rows else None

    trend = trend_series(members, date_cols, TREND_WINDOW, join_map)
    arrow = direction_arrow(trend)
    ratio = filled / total if total else 0.0

    # ---- 成员状态行（每名成员一个状态标记）----
    status_icons = {"filled": "✅", "leave": "🏖", "unfilled": "⚠️", "not_joined": "📅"}
    member_line = " · ".join(
        f"{p['name']}{status_icons[member_status(p, latest, exempt, join_map, cd)]}"
        for p in members)

    # ---- 考勤异常（本周期窗口，真实打卡）----
    anomalies = attendance_anomalies({p["name"] for p in members})
    if anomalies:
        shown = "、".join(
            f"{esc(n)}（{' · '.join(v)}）"
            for n, v in sorted(anomalies.items())[:4])
        if len(anomalies) > 4:
            shown += f" 等 {len(anomalies)} 人"
        att_line = f"\n<font color='red'>⏰ 考勤异常（{att_window}）：{shown}</font>"
    else:
        att_line = ""

    # ---- 逾期明细文本 ----
    # 来源优先级：台账（状态=待开始 且 截止已过 → 台账·超期未交付）→ 关键词自动扫描补充
    # → 人工核对档案。请假豁免人员的「未填报」条目经 _visible_overdue 过滤不展示
    # （请假即豁免，如董修奇 8.03–9.01 暑假——未填报确因在假，不作为逾期异常）；
    # 任务延期类条目（如高博的延期标记）不受影响；自动扫描仍跳过豁免/未入职人员。
    overdue_rows = _visible_overdue(known_overdue, exempt)
    overdue_rows += auto_overdue_extra(members, latest, known_overdue, exempt, join_map)
    if overdue_rows:
        lines = [f"**④ 逾期明细**（{len(overdue_rows)} 条）"]
        for owner, task, mark in overdue_rows:
            lines.append(
                f"<font color='red'>⚠️ {esc(owner)}｜{esc(task)} — "
                f"{esc(mark)}</font>")
        overdue_text = "\n".join(lines)
    else:
        overdue_text = "**④ 逾期明细**\n✅ 本周期无逾期"

    # ---- 请假标注（飞书审批真实数据；不附审批链接，普通成员无权限查看）----
    g_leaves = group_leaves([p["name"] for p in members], leave_window)
    if g_leaves:
        lv_lines = [f"**⑤ 请假标注**（飞书审批 · 近周期 {len(g_leaves)} 人）"]
        for e in g_leaves:
            days = f"{e['days']:g}" if e["days"] else ""
            span = f"{e['start'][5:10]} ~ {e['end'][5:10]}"
            lv_lines.append(
                f"🏖 {esc(e['name'])} · {esc(e['types'])} {span}"
                + (f"（{days}天）" if days else "") + " ✅已批准")
        leave_text = "\n".join(lv_lines)
    else:
        leave_text = "**⑤ 请假标注**\n本周期窗口内无已批准请假"

    trend_text = " → ".join(f"{f}/{t}（{lv}放假）" if lv else f"{f}/{t}"
                            for _, f, t, lv in trend)
    trend_line = (f"**⑦ 趋势**（近 {TREND_WINDOW} 期填报率 · 括号 N 放假 = "
                  f"当日请假在假，请假豁免填报）\n{trend_text} {arrow}")

    elements = [
        _md_div(
            f"**{esc(mentor)}** mentor · 组内 {total} 人 · 数据截至 {as_of}\n"
            f"成员：{member_line}"),
        {"tag": "hr"},
        _md_div(
            f"**① 人员概况**\n"
            f"填报 **{filled}/{total}** 人 · 请假 {leave} · 未按时 {unfilled}"
            + (f" ⚠️（{esc('、'.join(unfilled_names))}）" if unfilled else "")
            + (f" · 未入职 {not_joined} 📅" if not_joined else "")
            + att_line),
        _md_div(
            f"**② 目标数量**（{goals_src}口径）\n"
            f"大目标 {goals_b} ｜ 小目标 {goals_s} ｜ 本周期 TODO {goals_t}"),
        _md_div(
            f"**③ 完成情况**（{goals_src}）\n"
            f"✅ 已闭环 {closed} · 🔄 进行中 {doing}"
            + (f" · ⏳ 待开始 {waiting}" if waiting else "")
            + (f" · 🏖 顺延 {deferred}" if deferred else "")
            + f" · ⚠️ 逾期 {overdue_cnt} · 填报率 {round(ratio * 100)}%"),
        _md_div(overdue_text),
        _md_div(leave_text),
        _md_div(evaluation_text(evaluation, goals_src)),
        _md_div(trend_line),
        {"tag": "hr"},
        _md_div(_todo_source_line()),
        *(_md_div(
            f"📊 报表：[打开 {ledger_src.get('name', 'TODO 台账')}]({ledger_src['url']})")
          for _ in [0] if ledger_src.get("url")),
        _note(source_note(as_of, att_window)),
    ]

    return {
        "config": {"wide_screen_mode": True},
        "header": {
            "template": health_template(overdue_cnt, unfilled, ratio),
            "title": {"tag": "plain_text",
                      "content": f"📋 周期报表 · {latest} 第{len(date_cols)}周期 · {mentor}团队"},
        },
        "elements": elements,
    }


def build_boss_card(people: list[dict], latest: str, date_cols: list[str],
                    exempt: set[str], join_map: dict[str, str]) -> dict:
    """Boss 全公司总览卡：全局指标 + 团队维度 + 台账总览 + 逾期明细 + 请假 + 考勤 + 趋势 + 链接。"""
    by_mentor = defaultdict(list)
    for p in people:
        by_mentor[p["mentor"] or "(未分组)"].append(p)

    cd = _cycle_date(latest)
    leave_window, att_window = runtime_windows(latest)
    as_of = data_as_of()
    cal = load_calibration()

    def group_overdue(m: str) -> list[tuple[str, str, str]]:
        """某组逾期来源：该 mentor 台账（最新周期）→ 人工核对档案。"""
        ledger = load_ledger(m)
        rows = ledger_latest_rows(ledger)
        if rows:
            return overdue_from_ledger(rows)
        return calibration_overdue(cal, m)

    total = len(people)
    filled = sum(1 for p in people if member_status(p, latest, exempt, join_map, cd) == "filled")
    leave_names = {e["name"] for e in group_leaves([p["name"] for p in people], leave_window)}
    leave = len(leave_names)
    unfilled = sum(1 for p in people if member_status(p, latest, exempt, join_map, cd) == "unfilled")
    unfilled_names = [p["name"] for p in people
                      if member_status(p, latest, exempt, join_map, cd) == "unfilled"]
    not_joined = sum(1 for p in people
                     if member_status(p, latest, exempt, join_map, cd) == "not_joined")

    # 团队维度行：按组内人数降序（周熠 10 人最前），每行 人数/填报/未填/请假/逾期标记/填报率
    team_rows = []
    ledger_rows_by_mentor: dict[str, list[dict]] = {}
    for m, members in by_mentor.items():
        member_names = [p["name"] for p in members]
        f = sum(1 for p in members if member_status(p, latest, exempt, join_map, cd) == "filled")
        lv = len({e["name"] for e in group_leaves(member_names, leave_window)})
        uf = sum(1 for p in members if member_status(p, latest, exempt, join_map, cd) == "unfilled")
        od = len(group_overdue(m))
        pct = round(f / len(members) * 100) if members else 0
        flag = " ⚠️" if (uf or od) else ""
        team_rows.append(
            f"**{esc(m)}**｜{len(members)}人 · 填{f} · 未填{uf} · 假{lv} · "
            f"逾期{od} · {pct}%{flag}")
        ledger_rows_by_mentor[m] = ledger_latest_rows(load_ledger(m))
    team_rows.sort(key=lambda line: -int(re.search(r"｜(\d+)人", line).group(1)))
    team_text = "**团队维度（本周期）**\n" + "\n".join(team_rows)

    # 台账总览行（总览卡也要尽可能多地给出台账）：各组 ②目标数量/③完成情况/⑥评价 摘要。
    # 该组台账已注册（ledger_sources.json）且拉到数据 → 台账口径（附台账链接）；
    # 否则 → 人工核对档案口径（卡面标注核对时间，与 mentor 卡口径一致）。
    ledger_srcs = load_ledger_sources()
    ledger_lines = []
    for m in by_mentor:
        rows = ledger_rows_by_mentor.get(m) or []
        if rows:
            g = goal_counts_from_ledger(rows) or (0, 0, 0)
            c = closure_from_ledger(rows) or (0, 0, 0, 0, 0)
            ev = evaluation_from_ledger(rows)
            src = f"台账·截至{(load_ledger(m).get('latest_cycle') or '')[-5:] or '本周期'}"
            parts = [f"大{g[0]} 小{g[1]} TODO{g[2]}",
                     f"✅{c[0]} 🔄{c[1]}"
                     + (f" ⏳{c[2]}" if c[2] else "")
                     + (f" 🏖{c[3]}" if c[3] else "")
                     + f" ⚠️{c[4]}"]
            if ev:
                parts.append(f"★{ev['avg']}（{ev['n']}评）")
        else:
            g = calibration_goals(cal, m) or (0, 0, 0)
            c3 = calibration_closure(cal, m) or (0, 0, 0)
            src = f"人工核对 {cal.get('calibrated_at', '')[:10]}"
            parts = [f"大{g[0]} 小{g[1]} TODO{g[2]}",
                     f"✅{c3[0]} 🔄{c3[1]} ⚠️{c3[2]}"]
        line = f"**{esc(m)}**｜{' · '.join(parts)}（{src}）"
        lsrc = ledger_srcs.get(m, {})
        if lsrc.get("url"):
            line += f"  📊[台账]({lsrc['url']})"
        ledger_lines.append(line)
    ledger_text = "**📊 台账总览（各组 ②③⑥）**\n" + "\n".join(ledger_lines)

    # 全公司逾期明细：台账（状态=逾期，仅当 mentor 已使用）→ 自动扫描补充
    # → 人工核对档案。请假豁免人员的「未填报」条目不展示（如董修奇 8.03–9.01
    # 暑假，未填报确因在假）；任务延期类条目保留；自动扫描仍跳过豁免/未入职人员。
    all_overdue = []
    for m in by_mentor:
        known = group_overdue(m)
        for row in _visible_overdue(known, exempt):
            all_overdue.append((m, *row))
        all_overdue += [
            (m, *row) for row in auto_overdue_extra(by_mentor[m], latest, known, exempt, join_map)
        ]
    if all_overdue:
        od_lines = [f"**⚠️ 逾期明细（全公司 {len(all_overdue)} 条）**"]
        for group, owner, task, mark in all_overdue:
            od_lines.append(
                f"<font color='red'>⚠️ {esc(owner)}（{esc(group)}）｜"
                f"{esc(task)} — {esc(mark)}</font>")
        overdue_text = "\n".join(od_lines)
    else:
        overdue_text = "**⚠️ 逾期明细**\n✅ 全公司本周期无逾期"

    # 请假标注（真实审批；豁免未填报 + 已填报但请假中 都列；不附审批链接，普通成员无权限查看）
    g_leaves = group_leaves([p["name"] for p in people], leave_window)
    if g_leaves:
        leave_text = "**🏖 请假标注**（飞书审批 · 近周期 " + str(len(g_leaves)) + " 人）\n" + "\n".join(
            f"🏖 {esc(e['name'])}（{esc(by_mentor_name(people, e['name']))}）· "
            f"{esc(e['types'])} {e['start'][5:10]} ~ {e['end'][5:10]}"
            + (f"（{e['days']:g}天）" if e["days"] else "") + " ✅已批准"
            for e in sorted(g_leaves, key=lambda x: x["start"]))
    else:
        leave_text = "**🏖 请假标注**\n本周期窗口内无已批准请假"

    # 考勤异常（全公司）
    anomalies = attendance_anomalies({p["name"] for p in people})
    if anomalies:
        shown = "、".join(
            f"{esc(n)}（{' · '.join(v)}）"
            for n, v in sorted(anomalies.items())[:6])
        if len(anomalies) > 6:
            shown += f" 等 {len(anomalies)} 人"
        att_line = f"**⏰ 考勤异常**（{att_window} · 真实打卡）\n<font color='red'>{shown}</font>"
    else:
        att_line = f"**⏰ 考勤异常**（{att_window}）\n✅ 无迟到/早退/缺卡"

    trend = trend_series(people, date_cols, TREND_WINDOW, join_map)
    arrow = direction_arrow(trend)
    trend_text = " → ".join(f"{f}/{t}（{lv}放假）" if lv else f"{f}/{t}"
                            for _, f, t, lv in trend)

    ratio = filled / total if total else 0.0
    overdue_total = len(all_overdue)

    elements = [
        _md_div(
            f"在册 **{total}** 人 · **{len(by_mentor)}** 团队 · "
            f"本周期填报 **{filled}/{total}**（{round(ratio * 100)}%）· "
            f"请假 {leave} · 未按时 {unfilled}"
            + (f" · 未入职 {not_joined} 📅" if not_joined else "")
            + (f"\n⚠️ 未按时：{esc('、'.join(unfilled_names))}" if unfilled else "")),
        {"tag": "hr"},
        _md_div(team_text),
        _md_div(ledger_text),
        {"tag": "hr"},
        _md_div(overdue_text),
        _md_div(leave_text),
        _md_div(att_line),
        _md_div(f"**📈 趋势**（近 {TREND_WINDOW} 期全公司填报率 · 括号 N 放假 = "
                f"当日请假在假，请假豁免填报）\n{trend_text} {arrow}"),
        {"tag": "hr"},
        _md_div(_todo_source_line()),
        _note(source_note(as_of, att_window)),
    ]

    return {
        "config": {"wide_screen_mode": True},
        "header": {
            "template": "red" if (unfilled or overdue_total) else (
                "green" if ratio >= 1.0 else "blue"),
            "title": {"tag": "plain_text",
                      "content": f"📊 全公司 TODO 总览 · {latest} 第{len(date_cols)}周期"},
        },
        "elements": elements,
    }


def by_mentor_name(people: list[dict], name: str) -> str:
    for p in people:
        if p["name"] == name:
            return p["mentor"] or "(未分组)"
    return "?"


# --------------------------------------------------------------------------
# main
# --------------------------------------------------------------------------
def main() -> int:
    ap = argparse.ArgumentParser(description="生成 mentor 周期报表卡 + boss 总览卡")
    ap.add_argument("--xlsx", default="", help="可选：TODO LIST 导出 xlsx，重新解析后生成")
    args = ap.parse_args()

    date_cols, people = load_people(args.xlsx or None)
    if not date_cols:
        print("[err] 没有周期列（TODO LIST 数据源为空），无法生成", file=sys.stderr)
        return 1
    latest = date_cols[-1]            # 最新周期 = 数据源自动推导（不再维护 CYCLE 常量）
    cycle_no = len(date_cols)         # 周期序号 = 周期列数（历史列保留，新周期自动 +1）
    leave_window, att_window = runtime_windows(latest)
    as_of = data_as_of()
    ledger_sources = load_ledger_sources()
    registered = [m for m, s in ledger_sources.items() if s.get("app_token") and s.get("tables")]

    exempt = leave_exempt_names(leave_window)
    join_map = join_dates()
    print(f"[data] 最新周期列 {latest}（第 {cycle_no} 周期，共 {len(date_cols)} 列）")
    print(f"[data] 请假数据：{len(approved_leaves())} 条已批准，"
          f"窗口 {leave_window} 豁免 {len(exempt)} 人（{sorted(exempt)}）")
    print(f"[data] 考勤数据：{len(ATT_DATA.get('results', [])) if ATT_DATA else 0} 条"
          f"（窗口 {att_window}，截至 {as_of}）")
    print(f"[data] 入职数据：{len(join_map)} 人（本周期 {latest} 未入职 "
          f"{sum(1 for p in people if not joined_by(join_map, p['name'], _cycle_date(latest)))} 人）")
    print(f"[data] 已注册台账：{registered if registered else '无 → 全部回落人工核对口径'}")

    by_mentor = defaultdict(list)
    for p in people:
        by_mentor[p["mentor"] or "(未分组)"].append(p)
    # mentor 名单由数据派生（组名），open_id 由通讯录 roster 按名解析
    mentor_names = [m for m in by_mentor if m != "(未分组)"]
    oids = mentor_oids(mentor_names)

    # 防呆：本周期 TODO 数为 0 但数据源能解析出条目 → 必然漏数，告警
    cal = load_calibration()
    for mentor in mentor_names:
        members = by_mentor.get(mentor, [])
        if not members:
            continue
        ledger_rows = ledger_latest_rows(load_ledger(mentor))
        if ledger_rows:
            g = goal_counts_from_ledger(ledger_rows)
            manual_todo = g[2] if g else 0
        else:
            manual_todo = calibration_goals(cal, mentor)[2]
        auto = auto_todo_count(members, latest, join_map)
        if manual_todo == 0 and auto > 0:
            print(f"[warn] {mentor} 台账/人工档案 本周期 TODO=0，但数据源自动解析到 "
                  f"{auto} 条主条目，请复核台账 ledger 与人工核对档案！", file=sys.stderr)

    out: dict[str, dict] = {}
    for mentor in mentor_names:
        members = by_mentor.get(mentor, [])
        oid = oids.get(mentor, "")
        if not oid:
            print(f"[warn] mentor {mentor} 在通讯录 roster 中无 open_id，跳过发卡",
                  file=sys.stderr)
            continue
        card = build_mentor_card(mentor, members, latest, date_cols, exempt, join_map)
        out[mentor] = {"oid": oid, "card": card}

    out["__boss__"] = {"card": build_boss_card(people, latest, date_cols, exempt, join_map)}

    dest = Path(__file__).resolve().parent / "mentor_cards.json"
    dest.write_text(json.dumps(out, ensure_ascii=False), encoding="utf-8")

    print(f"[ok] 已生成 {len(out) - 1} 张 mentor 卡 + 1 张 boss 卡 → {dest}")
    for mentor, entry in out.items():
        if mentor == "__boss__":
            print(f"  boss 卡 ✅（template={entry['card']['header']['template']}）")
        else:
            members = by_mentor.get(mentor, [])
            f = sum(1 for p in members
                    if member_status(p, latest, exempt, join_map, _cycle_date(latest)) in ("filled", "leave"))
            print(f"  {mentor}（{len(members)}人 填{f}）✅ "
                  f"template={entry['card']['header']['template']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
if __name__ == "__main__":
    raise SystemExit(main())
