"""Convert 国央企领导周报Agent-全量问题清单_93问.xlsx into a baseline answer set.

Why a converter and not a hand-written jsonl: the xlsx is Guoshu's own
deliverable and will be revised.  Re-running this script re-derives the answer
set; transcribing it by hand would fork the two.

What this set is, and how it differs from the 396-question one:

* Its reference answers are *prose written by the business side*, not SQL result
  sets.  There is no ``gold_sql`` to audit, so the usual审计 (missing gates,
  COUNT(*) where a task count was asked) does not apply -- instead every number
  in the 参考答案 was spot-checked against the mock store before adopting this
  set as a standard, and the checks passed (A01 128/82/46, B01 119, B08 18,
  C01 117 under the formal gate).
* It is anchored at **2026-08-17**, two days later than the 396-question set.
  Only 08-17 reproduces B06's 14/28/4 buckets; 08-15 gives 17/28/1.  Callers
  must therefore run the mock service with ``GUOSHU_AS_OF=2026-08-17``.
* Only 40 of the 93 are plain data lookups.  33 are 规则信号 (the answer is a
  judgement plus a recommended action) and 20 are 暂不可答 (the answer is a
  refusal naming the missing fields).  A fact-equivalence grader would mark a
  correct refusal as wrong, which is why ``grade_mode`` travels with each row.

Usage:
    python tests/build_g93_answers.py \
        --xlsx "~/Desktop/国数周报/国央企领导周报Agent-全量问题清单_93问.xlsx" \
        --out tests/g93-answers.jsonl
"""

# ruff: noqa: RUF001  中文字段名与口径文案里的全角标点是数据, 不能换成半角。
# ruff: noqa: T201  这是命令行脚本, stdout 就是它的输出通道。
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

SHEET = "问题明细"
HEADER_KEY = "问题编号"

STATUS_TO_MODE = {
    "数据事实": "fact",
    "规则信号": "signal",
    "暂不可答": "refusal",
}
"""回答状态 -> grading rubric.

``fact`` grades numbers and item sets.  ``signal`` grades whether the判断 and
its supporting numbers hold, allowing the wording of the recommendation to
differ.  ``refusal`` grades whether the agent declined *and* named the missing
data -- an agent that guesses a number here has failed even if the number is
plausible.
"""

# 交互形式 -> the kind bucket baseline.py reports by.  Kept coarse on purpose:
# 指标卡/表格/文本 says how the answer is rendered, and the only distinction
# that changes grading is single-value vs set vs prose.
FORM_TO_KIND = {
    "指标卡": "scalar",
    "表格": "table",
    "文本": "prose",
    "周报预览": "prose",
    "文件下载": "prose",
}


def _load_rows(xlsx: Path) -> list[dict[str, str]]:
    # Imported here, not at module top: openpyxl is only needed to re-derive the
    # answer set from Guoshu's xlsx.  A baseline run reads the generated jsonl and
    # must not require the dependency.
    try:
        import openpyxl  # noqa: PLC0415
    except ModuleNotFoundError:  # pragma: no cover - environment guard
        print("需要 openpyxl：pip install openpyxl", file=sys.stderr)
        raise SystemExit(2) from None

    book = openpyxl.load_workbook(xlsx, data_only=True)
    if SHEET not in book.sheetnames:
        raise SystemExit(f"{xlsx} 里没有工作表 {SHEET}，只有 {book.sheetnames}")
    sheet = book[SHEET]
    raw = list(sheet.iter_rows(values_only=True))

    header_at = next((i for i, row in enumerate(raw) if row and row[0] == HEADER_KEY), -1)
    if header_at < 0:
        raise SystemExit(f"{SHEET} 里找不到表头行（首列 {HEADER_KEY}）")
    header = [("" if c is None else str(c).strip()) for c in raw[header_at]]

    out: list[dict[str, str]] = []
    for row in raw[header_at + 1 :]:
        if not row or not row[0]:
            continue
        out.append({header[i]: ("" if row[i] is None else str(row[i]).strip()) for i in range(len(header))})
    return out


def _difficulty(rec: dict[str, str]) -> str:
    """Assign difficulty from what the question demands, not from a guess.

    Deliberately mechanical so the per-difficulty rates in the report mean
    something stable: refusals and multi-source signals are the hard end because
    they require the agent to reason about what the store *cannot* answer.
    """
    status = rec["回答状态"]
    if status == "暂不可答":
        return "expert"
    if status == "规则信号":
        return "hard"
    if "、" in rec.get("数据依据", "") or "综合" in rec.get("数据依据", ""):
        return "medium"
    return "easy"


def convert(rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for rec in rows:
        status = rec["回答状态"]
        mode = STATUS_TO_MODE.get(status)
        if mode is None:
            raise SystemExit(f"{rec[HEADER_KEY]} 的回答状态无法识别：{status!r}")
        qid = rec[HEADER_KEY]
        records.append(
            {
                "id": qid,
                "type_id": qid[:1],
                # Scene is the axis Guoshu reviews by, so it leads the category
                # label; the letter keeps per-block rates readable in the report.
                "category": f"{qid[:1]} {rec['领导场景']}",
                "type": rec["原始主题"],
                "difficulty": _difficulty(rec),
                "question": rec["问题原文"],
                "kind": FORM_TO_KIND.get(rec["交互形式"], "prose"),
                "answer_status": status,
                "grade_mode": mode,
                "gold_answer": rec["参考答案"],
                "evidence": rec["数据依据"],
                "boundary": rec["回答边界"],
                "role": rec["主适用角色"],
                "granularity": rec["对象粒度"],
                "time_scope": rec["时间粒度"],
                "permission_note": rec["权限与脱敏要求"],
                "keywords": [k for k in rec["关键词"].replace("、", ",").split(",") if k],
            }
        )
    return records


def main() -> int:
    parser = argparse.ArgumentParser(description="93 问清单 -> baseline 答案集")
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--out", default=str(Path(__file__).parent / "g93-answers.jsonl"))
    args = parser.parse_args()

    xlsx = Path(args.xlsx).expanduser()
    if not xlsx.exists():
        print(f"清单不存在：{xlsx}", file=sys.stderr)
        return 2

    records = convert(_load_rows(xlsx))
    out = Path(args.out)
    out.write_text(
        "".join(json.dumps(r, ensure_ascii=False) + "\n" for r in records),
        encoding="utf-8",
    )

    by_mode: dict[str, int] = {}
    for r in records:
        by_mode[r["grade_mode"]] = by_mode.get(r["grade_mode"], 0) + 1
    print(f"写入 {out}：{len(records)} 题")
    for mode, n in sorted(by_mode.items()):
        print(f"  {mode:<8} {n:>3} 题")
    print("提醒：本套锚定 2026-08-17，跑之前要 GUOSHU_AS_OF=2026-08-17 起 mock 服务")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
