"""Write-Excel tool - create real .xlsx spreadsheets instead of markdown tables."""

from __future__ import annotations

import json
from typing import Any

import anyio


def _build_workbook(file_path: str, rows: list[list[Any]], sheet_name: str, header: bool) -> int:
    """Build and save an .xlsx workbook synchronously. Returns the row count written."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Sheet1"  # Excel caps sheet titles at 31 chars

    col_widths: dict[int, int] = {}
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if header and r_idx == 1:
                cell.font = Font(bold=True)
            text_len = len(str(value)) if value is not None else 0
            if text_len > col_widths.get(c_idx, 0):
                col_widths[c_idx] = text_len

    for c_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max(width + 2, 8), 60)

    wb.save(file_path)
    return len(rows)


async def write_excel(file_path: str, rows_json: str, sheet_name: str = "Sheet1", header: bool = True) -> str:
    """Create a real Excel (.xlsx) file from tabular data.

    Use this instead of writing a markdown table when the user asks for a
    table, spreadsheet, or Excel file. Each inner list is one row; cells may
    be strings or numbers.

    Args:
        file_path: Output path for the .xlsx file (e.g. "report.xlsx").
        rows_json: JSON-encoded 2D array of rows, e.g.
            '[["Name", "Score"], ["Alice", 92], ["Bob", 88]]'.
        sheet_name: Worksheet name (truncated to Excel's 31-char limit).
        header: When True, bold the first row as a header.

    Returns:
        Success message with the row count, or an error message.
    """
    if not file_path.lower().endswith(".xlsx"):
        file_path = f"{file_path}.xlsx"

    try:
        rows = json.loads(rows_json)
    except json.JSONDecodeError as e:
        return f"[Error] rows_json is not valid JSON: {e}"

    if not isinstance(rows, list) or not all(isinstance(r, list) for r in rows):
        return "[Error] rows_json must be a 2D array, e.g. [[\"A\", \"B\"], [1, 2]]"
    if not rows:
        return "[Error] rows_json is empty; provide at least one row"

    path = anyio.Path(file_path)
    parent = path.parent
    if not await parent.exists():
        await parent.mkdir(parents=True, exist_ok=True)

    try:
        count = await anyio.to_thread.run_sync(_build_workbook, file_path, rows, sheet_name, header)
    except Exception as e:  # openpyxl raises assorted errors on bad values
        return f"[Error] Failed to write Excel file: {e!r}"

    return f"[OK] Wrote {count} row(s) to {file_path}"
