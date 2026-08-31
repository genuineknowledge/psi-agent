"""P1-3 conversation-history export (Excel / PDF).

Exports the *conversation* — questions and answers — as opposed to the P1-1
weekly summary, which is generated material. The Gateway keeps the history;
the BFF fetches it and renders the two formats.
"""

from __future__ import annotations

import io
import math
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

_HEADER_FILL = PatternFill("solid", fgColor="0F6B54")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_WRAP = Alignment(wrap_text=True, vertical="top")


INJECTION_MARKER = "【本次回答要求】"
INJECTION_SEPARATOR = "\n\n问题:"


def _strip_injection_prefix(text: str) -> str:
    """The BFF prefixes forwarded questions with an answer-organisation
    instruction, which the Gateway records verbatim — exports must show the
    question the user actually asked."""
    if text.startswith(INJECTION_MARKER) and INJECTION_SEPARATOR in text:
        return text.split(INJECTION_SEPARATOR, 1)[1]
    return text


def _pair_rounds(messages: list[dict[str, Any]]) -> list[tuple[str, str]]:
    """Fold the transcript into question/answer rounds.

    A round is one user message plus the assistant messages that follow it —
    that is how a conversation reads, and it keeps exports free of the
    "every message is a row" noise.
    """
    rounds: list[tuple[str, str]] = []
    question: str | None = None
    answers: list[str] = []
    for message in messages:
        if message.get("role") == "user":
            if question is not None:
                rounds.append((question, "\n\n".join(answers)))
            question = _strip_injection_prefix(str(message.get("text", "")))
            answers = []
        else:
            answers.append(str(message.get("text", "")))
    if question is not None:
        rounds.append((question, "\n\n".join(answers)))
    return rounds


def _plain_text(markdown: str) -> str:
    """Strip light markdown decorators so exports read as plain text:
    **bold**, `code`, heading hashes and quote markers. Table pipes stay —
    they read as column separators in a plain-text cell."""
    text = markdown.replace("**", "").replace("`", "")
    text = re.sub(r"^#{1,6}\s*", "", text, flags=re.MULTILINE)
    text = re.sub(r"^>\s?", "", text, flags=re.MULTILINE)
    return text


_LINE_HEIGHT = 14.5
_ROW_PAD = 8.0
_MAX_ROW_HEIGHT = 400.0  # Excel hard cap is 409pt; leave a little headroom


def _estimate_lines(text: str, chars_per_line: int) -> int:
    """Lines a text needs in a wrap-text cell of ~chars_per_line CJK chars."""
    total = 0
    for raw in text.split("\n"):
        total += max(1, math.ceil(len(raw) / chars_per_line))
    return total


def build_excel(messages: list[dict[str, Any]]) -> bytes:
    """One row per round: index, question, answer.

    Two things make this readable in Excel/WPS: the markdown is reduced to
    plain text, and each row's height is sized from its content — wrap_text
    alone does not raise the row, so a long answer used to open as a single
    clipped line.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "对话历史"
    for cell, header in zip(["A1", "B1", "C1"], ["序号", "问题", "回答"], strict=True):
        sheet[cell] = header
        sheet[cell].fill = _HEADER_FILL
        sheet[cell].font = _HEADER_FONT
    for index, (question, answer) in enumerate(_pair_rounds(messages), start=1):
        question = _plain_text(question)
        answer = _plain_text(answer)
        row = index + 1
        sheet.cell(row=row, column=1, value=index)
        q_cell = sheet.cell(row=row, column=2, value=question)
        q_cell.alignment = _WRAP
        a_cell = sheet.cell(row=row, column=3, value=answer)
        a_cell.alignment = _WRAP
        # Column widths 52/96 ≈ 26/48 CJK characters per line.
        height = max(_estimate_lines(question, 26), _estimate_lines(answer, 48)) * _LINE_HEIGHT + _ROW_PAD
        sheet.row_dimensions[row].height = min(_MAX_ROW_HEIGHT, height)
    sheet.column_dimensions["A"].width = 8
    sheet.column_dimensions["B"].width = 52
    sheet.column_dimensions["C"].width = 96
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))  # built into every PDF reader, no font file needed

_USABLE_MM = 210 - 18 * 2  # A4 width minus the document's left/right margins


def _escape_xml(text: str) -> str:
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _inline(text: str) -> str:
    """Inline markdown → reportlab markup: **bold**, `code`. Text is XML-
    escaped first so the markdown markers cannot smuggle tags."""
    text = _escape_xml(text)
    text = re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", text)
    text = re.sub(r"`([^`]+)`", r'<font face="Courier" size="9">\1</font>', text)
    return text


def _split_row(line: str) -> list[str]:
    text = line.strip()
    if text.startswith("|"):
        text = text[1:]
    if text.endswith("|"):
        text = text[:-1]
    return [cell.strip() for cell in text.split("|")]


def _is_table_separator(line: str) -> bool:
    return bool(re.match(r"^\|?[\s:|-]+\|?$", line.strip())) and "-" in line


def _pdf_table(rows: list[list[str]]) -> Table:
    """One markdown table → a bordered reportlab table with a green header."""
    n_cols = max(len(row) for row in rows)
    cell_style = ParagraphStyle("cell", fontName="STSong-Light", fontSize=9, leading=12)
    data: list[list[Paragraph]] = []
    for row in rows:
        data.append(
            [Paragraph(_inline(cell).replace("\n", "<br/>"), cell_style) for cell in row + [""] * (n_cols - len(row))]
        )
    table = Table(data, colWidths=[_USABLE_MM / n_cols * mm] * n_cols, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F6B54")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D5DED9")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    return table


def _pdf_story_from_markdown(markdown: str) -> list[Any]:
    """Light markdown → PDF blocks: headings, tables (native), lists, and
    paragraphs with **bold** / `code` inline markup."""
    body_style = ParagraphStyle("body", fontName="STSong-Light", fontSize=10.5, leading=15, spaceAfter=6)
    list_style = ParagraphStyle("list", parent=body_style, leftIndent=14)
    code_style = ParagraphStyle(
        "code",
        fontName="Courier",
        fontSize=9,
        leading=12,
        backColor=colors.HexColor("#F2F4F3"),
    )
    heading_sizes = {1: 14, 2: 12.5, 3: 11.5}

    lines = markdown.split("\n")
    story: list[Any] = []
    index = 0
    while index < len(lines):
        stripped = lines[index].strip()
        if not stripped:
            index += 1
            continue
        # Native table: header row + separator + body rows.
        if stripped.startswith("|") and index + 1 < len(lines) and _is_table_separator(lines[index + 1]):
            rows = [_split_row(stripped)]
            index += 2
            while index < len(lines) and lines[index].strip().startswith("|"):
                rows.append(_split_row(lines[index]))
                index += 1
            story.append(_pdf_table(rows))
            story.append(Spacer(1, 2 * mm))
            continue
        # Heading.
        heading = re.match(r"^(#{1,6})\s+(.*)", stripped)
        if heading:
            level = len(heading.group(1))
            style = ParagraphStyle(
                f"h{level}",
                fontName="STSong-Light",
                fontSize=heading_sizes.get(level, 11),
                leading=heading_sizes.get(level, 11) + 4,
                spaceBefore=6,
                spaceAfter=4,
            )
            story.append(Paragraph(_inline(heading.group(2)), style))
            index += 1
            continue
        # List item.
        if re.match(r"^[-*]\s+", stripped) or re.match(r"^\d+[.)]\s+", stripped):
            item = re.sub(r"^[-*]\s+", "", stripped)
            item = re.sub(r"^\d+[.)]\s+", "", item)
            story.append(Paragraph(f"• {_inline(item)}", list_style))
            index += 1
            continue
        # Fenced code block (rare in answers; render monospaced).
        if stripped.startswith("```"):
            code_lines: list[str] = []
            index += 1
            while index < len(lines) and not lines[index].strip().startswith("```"):
                code_lines.append(lines[index])
                index += 1
            index += 1  # closing fence
            story.append(Paragraph(_escape_xml("\n".join(code_lines)).replace("\n", "<br/>"), code_style))
            story.append(Spacer(1, 2 * mm))
            continue
        # Paragraph: run until a blank line or the next block start.
        buffer_lines = [stripped]
        index += 1
        while (
            index < len(lines)
            and lines[index].strip()
            and not (
                lines[index].strip().startswith("|")
                or lines[index].strip().startswith("#")
                or lines[index].strip().startswith("```")
                or re.match(r"^[-*]\s+", lines[index].strip())
                or re.match(r"^\d+[.)]\s+", lines[index].strip())
            )
        ):
            buffer_lines.append(lines[index].strip())
            index += 1
        story.append(Paragraph(_inline(" ".join(buffer_lines)), body_style))
    return story


def build_pdf(messages: list[dict[str, Any]]) -> bytes:
    """Readable PDF: title, then one block per question/answer round with
    answers rendered from their markdown (tables as real tables, bold and
    code inline, headings and lists structured)."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
        title="对话历史",
    )
    title_style = ParagraphStyle("title", fontName="STSong-Light", fontSize=16, spaceAfter=14)
    question_style = ParagraphStyle(
        "question", fontName="STSong-Light", fontSize=11, textColor=colors.HexColor("#0F6B54"), spaceAfter=4
    )

    story: list[Any] = [Paragraph("对话历史", title_style)]
    for index, (question, answer) in enumerate(_pair_rounds(messages), start=1):
        story.append(Paragraph(f"{index}. 问:{_escape_xml(question)}", question_style))
        story.extend(_pdf_story_from_markdown(answer))
        story.append(Spacer(1, 2 * mm))
    doc.build(story)
    return buffer.getvalue()
